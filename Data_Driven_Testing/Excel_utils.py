import openpyxl


def get_row_count(file, sheet_name):
    #counts the number of rows in the table in the given file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row

def get_column_count(file, sheet_name):
    #counts the number of columns in the table in the given file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column

def read_data(file, sheet_name, rownum, columnnum):
    #reads data from the file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=rownum, column=columnnum).value

def write_data(file, sheet_name, rownum, columnnum, data):
    #writes data to the file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)


